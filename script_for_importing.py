import pandas as pd
import re
import os
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook


excel_files = [f for f in os.listdir('.') if f.endswith('.xlsx') or f.endswith('.xls')]
if not excel_files:
    raise FileNotFoundError("No Excel files found in the current directory.")

elif len(excel_files) == 1:
    file_path = excel_files[0]
    print(f"Using Excel file: {file_path}")
else:
    print("Multiple Excel files found:")
    for i, f in enumerate(excel_files, start=1):
        print(f"{i}: {f}")
    choice = input("Enter the number of the file you want to use: ")
    try:
        file_path = excel_files[int(choice)-1]
    except:
        raise ValueError("Invalid selection.")

xls = pd.ExcelFile(file_path)


# Load the 'Eigenaren' sheet
df = pd.read_excel(xls, sheet_name='Eigenaren')
if 'Eigenaren' not in xls.sheet_names:
    raise ValueError(f"The sheet 'Eigenaren' was not found in {file_path}. Available sheets: {xls.sheet_names}")

def parse_name(full_name):
    if pd.isna(full_name):
        return "", "", "", ""

    full_name = str(full_name).strip()

    
    title_map = {
        "de heer": "Man", "dhr.": "Man", "dhr": "Man", "heer": "Man",
        "mevrouw": "Vrouw", "mw.": "Vrouw", "mw": "Vrouw"
    }
    title = ""
    for k, v in title_map.items():
        if full_name.lower().startswith(k):
            title = v
            full_name = full_name[len(k):].strip()
            break

    # Extract voorletters (in or out of parentheses)
    voorletters_match = re.match(r"\(?([A-Z][a-zA-Z.]*)\)?", full_name)
    if voorletters_match:
        voorletters = voorletters_match.group(1).strip()
        full_name = full_name[voorletters_match.end():].strip()
    else:
        voorletters = ""

    
    tussenvoegsel_set = {
        "van", "van de", "van der", "de", "den", "ter", "ten", "het", "der", "op", "aan", "in", "uit"
    }

    
    name_parts = full_name.split()
    tussenvoegsel = ""
    achternaam_parts = []

    i = 0
    while i < len(name_parts):
        candidate = name_parts[i].lower()
        #
        if i < len(name_parts) - 1 and f"{candidate} {name_parts[i + 1].lower()}" in tussenvoegsel_set:
            tussenvoegsel = f"{candidate} {name_parts[i + 1]}"
            i += 2
            break
        elif candidate in tussenvoegsel_set:
            tussenvoegsel = candidate
            i += 1
            break
        else:
            break  

    achternaam_parts = name_parts[i:]
    achternaam = " ".join(achternaam_parts).strip()

    return title, voorletters.strip(), tussenvoegsel.strip(), achternaam.strip()


def parse_address(address):
    
    if pd.isna(address):
        return None, None, None, None

    try:
        street_part, rest = address.split(',', 1)
        postcode_match = re.search(r'\d{4}\s?[A-Z]{2}', rest)
        postcode = postcode_match.group(0) if postcode_match else ""
        city = rest.replace(postcode, '').strip() if postcode else rest.strip()

        street_name = re.sub(r'\d.*', '', street_part).strip()
        house_number = street_part.replace(street_name, '').strip()

        return street_name, house_number, postcode.strip(), city.strip()
    except Exception:
        return None, None, None, None


# Group rows by Index nr. to collect cohabitants
grouped = df.groupby('Index nr.')

# Define the target structure
output_columns = [
    "(Achter-) naam*", "Voorletters / -naam", "Tussenvoegsel", "Geslacht / type*",
    "Straatnaam*", "Huisnummer*", "Postcode*", "Plaats*",
    "Straatnaam postadres", "Huisnummer postadres", "Postcode postadres", "Plaats postadres",
    "Categorie", "Telefoonnummer 1", "Type telefoonnummer 1",
    "Telefoonnummer 2", "Type telefoonnummer 2", "Telefoonnummer 3", "Type telefoonnummer 3",
    "E-mailadressen", "IBAN", "Is debiteur", "Is crediteur", "Incassomachtiging afgegeven",
    "Factuur gewenst", "Factuurtoelichting gewenst",
    "Achternaam contactpersoon 1", "Voorletters contactpersoon 1", "Tussenvoegsel contactpersoon 1",
    "Geslacht contactpersoon 1", "Telefoonnummer 1 contactpersoon 1", "Telefoonnummer 2 contactpersoon 1",
    "E-mailadres contactpersoon 1",
    "Achternaam contactpersoon 2", "Voorletters contactpersoon 2", "Tussenvoegsel contactpersoon 2",
    "Geslacht contactpersoon 2", "Telefoonnummer 1 contactpersoon 2", "Telefoonnummer 2 contactpersoon 2",
    "E-mailadres contactpersoon 2"
]
output_data = []

for index, group in grouped:
    row = dict.fromkeys(output_columns, "")

    owners = group.reset_index(drop=True)
    people = []

    # Parse each person in the group
    for i, person in owners.iterrows():
        title, voorletters, tussenvoegsel, achternaam = parse_name(person["Eigenaar"])
        people.append({
            "title": title,
            "voorletters": voorletters,
            "tussenvoegsel": tussenvoegsel,
            "achternaam": achternaam,
            "email": person["Email eigenaar"],
            "telefoon": str(person["Telefoon eigenaar"]).split(',')[0].split(';')[0],
        })

    # Combine last names only (for hoofdinschrijving)
    combined_achternamen = " & ".join([p["achternaam"] for p in people])
    row["(Achter-) naam*"] = combined_achternamen

    # Main columns: leave other 3 empty
    row["Voorletters / -naam"] = ""
    row["Tussenvoegsel"] = ""
    row["Geslacht / type*"] = ""

    # Parse main address
    straat, huisnr, postcode, plaats = parse_address(owners.iloc[0]["Adres"])
    row["Straatnaam*"] = straat
    row["Huisnummer*"] = huisnr
    row["Postcode*"] = postcode
    row["Plaats*"] = plaats

    # Parse post address
    postadres = owners.iloc[0]["Postadres eigenenaar"]
    straat_post, huisnr_post, postcode_post, plaats_post = parse_address(postadres)
    row["Straatnaam postadres"] = straat_post
    row["Huisnummer postadres"] = huisnr_post
    row["Postcode postadres"] = postcode_post
    row["Plaats postadres"] = plaats_post

    # Other fields
    row["Categorie"] = owners.iloc[0]["Unittype"]
    row["E-mailadressen"] = people[0]["email"]
    row["Telefoonnummer 1"] = people[0]["telefoon"]

    # Contactpersoon 1 = owner 0
    if len(people) > 0:
        p1 = people[0]
        row["Achternaam contactpersoon 1"] = p1["achternaam"]
        row["Tussenvoegsel contactpersoon 1"] = p1["tussenvoegsel"]
        row["Voorletters contactpersoon 1"] = p1["voorletters"]
        row["Geslacht contactpersoon 1"] = p1["title"]
        row["E-mailadres contactpersoon 1"] = p1["email"]
        row["Telefoonnummer 1 contactpersoon 1"] = p1["telefoon"]

    # Contactpersoon 2 = owner 1 (if exists)
    if len(people) > 1:
        p2 = people[1]
        row["Achternaam contactpersoon 2"] = p2["achternaam"]
        row["Tussenvoegsel contactpersoon 2"] = p2["tussenvoegsel"]
        row["Voorletters contactpersoon 2"] = p2["voorletters"]
        row["Geslacht contactpersoon 2"] = p2["title"]
        row["E-mailadres contactpersoon 2"] = p2["email"]
        row["Telefoonnummer 1 contactpersoon 2"] = p2["telefoon"]

    output_data.append(row)

output_df = pd.DataFrame(output_data)
base_name = os.path.splitext(os.path.basename(file_path))[0]
output_filename = f"converted_{base_name}.xlsx"

# Save to Excel
output_df.to_excel(output_filename, index=False)

# Load and adjust column widths
wb = load_workbook(output_filename)
ws = wb.active

for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[column].width = max_length + 2

wb.save(output_filename)
print(f"Data has been processed and saved to '{output_filename}'.")